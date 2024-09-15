import json
import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

def get_one_seq_from_json(json_dict, seq_name, dataset_name, skip_bitrate=False):
    avg_psnrs = []
    avg_bpps = []
    avg_msssims = []
    
    continue_flag = False
    if skip_bitrate:
        continue_flag = True
    for result_dict in json_dict[dataset_name][seq_name].values():
        if continue_flag:
            continue_flag = False
            continue
        avg_psnrs.append(float(result_dict["ave_all_frame_psnr"]))
        avg_bpps.append(float(result_dict["ave_all_frame_bpp"]))
        avg_msssims.append(float(result_dict["ave_all_frame_msssim"]))
    
    return {"avg_psnrs":np.array(avg_psnrs), "avg_bpps":np.array(avg_bpps), "avg_msssims":np.array(avg_msssims)}
    
def get_one_seq_from_excel(sheet, seq_idx, bitrate_num):
    avg_psnrs = []
    avg_bpps = []
    avg_msssims = []
    idx = 0
    for i in sheet.iter_cols(min_row=2+seq_idx*bitrate_num, max_row=1+seq_idx*bitrate_num+bitrate_num, min_col=2,max_col=4):
        for j in i:
            if idx == 0:
                avg_bpps.append(float(j.value))
            elif idx == 1:
                avg_psnrs.append(float(j.value))
            elif idx == 2:
                if not j.value:
                    avg_msssims.append(float(0))
                else:
                    avg_msssims.append(float(j.value))
        idx += 1
    return {"avg_psnrs":np.array(avg_psnrs), "avg_bpps":np.array(avg_bpps), "avg_msssims":np.array(avg_msssims)}
    

seq_names = ['USTC_Badminton', 'USTC_BasketballDrill', 'USTC_BasketballPass', 'USTC_BicycleDriving', 'USTC_Dancing', 'USTC_FourPeople', 'USTC_ParkWalking', 'USTC_Running', 'USTC_ShakingHands', 'USTC_Snooker']
deep_codec_names = ['DVC-pro', 'DCVC', 'CANF', 'DCVC-TCM', 'DCVC-HEM', 'DCVC-OOFE', 'VNVC', 'SDD', 'DCVC-DC', 'DCVC-FM']
traditional_codec_names = ['HM', 'VTM']
codec_names = traditional_codec_names + deep_codec_names
two_single_model_codecs = ['DVC-pro', 'DCVC', 'DCVC-TCM', 'DCVC-HEM', 'VNVC', 'SDD', 'DCVC-DC']
codec_names_t = ['H.265/HEVC', 'H.266/VVC', 'DVC_Pro', 'DCVC', 'CANF-VC', 'TCM-VC', 'DCVC-HEM', 'OOFE', 'VNVC', 'SDD', 'DCVC-DC', 'DCVC-FM']
bitrate_nums = [5, 5, 4, 4, 4, 4, 4, 4, 4, 4, 4, 3]
bo_sets = ['DCVC-FM', 'DCVC-OOFE', 'EEM']
seq_idx = 0
all_seq_bpps = dict([(key,np.array([0.0 for idx in range(bitrate_num)])) for key, bitrate_num in zip(codec_names, bitrate_nums)])
all_seq_psnrs = dict([(key,np.array([0.0 for idx in range(bitrate_num)])) for key, bitrate_num in zip(codec_names, bitrate_nums)])
all_seq_msssims = dict([(key,np.array([0.0 for idx in range(bitrate_num)])) for key, bitrate_num in zip(codec_names, bitrate_nums)])
all_seq_msssim_bpps = dict([(key,np.array([0.0 for idx in range(bitrate_num)])) for key, bitrate_num in zip(codec_names, bitrate_nums)])
json_dir = './jsons/PSNR_models'
json_msssim_dir = './jsons/MSSSIM_models'
excel_dir = './excels'
fontdict = {'family' : 'Times New Roman', 'size'   : 16}

for seq_name in seq_names:
    ls = '-'
    marker = 'o'
    codec_idx = 0
    # for excel files
    for codec_name, bitrate_num in zip(traditional_codec_names, bitrate_nums[0:4]):
        if codec_name == 'AV1':
            excel_file_name = os.path.join(excel_dir, f'{codec_name}_444_anchor_nofirst.xlsx')
        else:
            excel_file_name = os.path.join(excel_dir, f'{codec_name}_444_anchor.xlsx')
        workbook = load_workbook(filename=excel_file_name)
        sheet = workbook['Sheet']
        
        curr_seq_codec_result = get_one_seq_from_excel(sheet, seq_idx, bitrate_num=bitrate_num)
        plt.figure(1)
        plt.plot(curr_seq_codec_result['avg_bpps'],curr_seq_codec_result['avg_psnrs'],label=codec_names_t[codec_idx], marker=marker, ls=ls)
        plt.figure(2)
        plt.plot(curr_seq_codec_result['avg_bpps'],curr_seq_codec_result['avg_msssims'],label=codec_names_t[codec_idx], marker=marker, ls=ls)
        
        all_seq_bpps[codec_name] += curr_seq_codec_result['avg_bpps'] / 10
        all_seq_psnrs[codec_name] += curr_seq_codec_result['avg_psnrs'] / 10
        all_seq_msssims[codec_name] += curr_seq_codec_result['avg_msssims'] / 10
        all_seq_msssim_bpps[codec_name] += curr_seq_codec_result['avg_bpps'] / 10
        
        codec_idx += 1
    
    # for json files
    for codec_name in deep_codec_names:
        if codec_idx == 10:
            ls = '-.'
            marker = '^'
        json_file_name = os.path.join(json_dir, f'{codec_name}.json')
        json_msssim_file_name = os.path.join(json_msssim_dir, f'msssim-{codec_name}.json')
        if codec_name in bo_sets:
            dataset_name = 'USTC-TD'
        else:
            dataset_name = 'iVC'
            
        with open(json_file_name) as f:
            json_dict = json.load(f)

        with open(json_msssim_file_name) as f_ms:
            json_dict_msssim = json.load(f_ms)
            
        if codec_name == 'DCVC-FM':
            curr_seq_codec_result = get_one_seq_from_json(json_dict, seq_name, dataset_name, skip_bitrate=True)
        else:
            curr_seq_codec_result = get_one_seq_from_json(json_dict, seq_name, dataset_name)
        
        if codec_name in two_single_model_codecs:
            curr_seq_msssim_codec_result = get_one_seq_from_json(json_dict_msssim, seq_name, dataset_name)
        else:
            curr_seq_msssim_codec_result = curr_seq_codec_result
        
            
        plt.figure(1)
        plt.plot(curr_seq_codec_result['avg_bpps'],curr_seq_codec_result['avg_psnrs'],label=codec_names_t[codec_idx], marker=marker, ls=ls)
        plt.figure(2)
        plt.plot(curr_seq_msssim_codec_result['avg_bpps'],curr_seq_msssim_codec_result['avg_msssims'],label=codec_names_t[codec_idx], marker=marker, ls=ls)
        
        all_seq_bpps[codec_name] += curr_seq_codec_result['avg_bpps'] / 10
        all_seq_psnrs[codec_name] += curr_seq_codec_result['avg_psnrs'] / 10
        all_seq_msssims[codec_name] += curr_seq_msssim_codec_result['avg_msssims'] / 10
        all_seq_msssim_bpps[codec_name] += curr_seq_msssim_codec_result['avg_bpps'] / 10
            
        codec_idx += 1
        
    plt.figure(1)
    plt.legend(loc='best')
    plt.subplots_adjust(left=0.12, right=0.995, top=0.990, bottom=0.105)
    # plt.title(f"{seq_name.replace('_', '-')}", fontdict=fontdict)
    plt.ylabel('PSNR (dB)', fontdict=fontdict)
    plt.xlabel('bits per pixel (bpp)', fontdict=fontdict)
    plt.grid()
    plt.savefig(f'./figs/PSNR_{seq_name}.png', dpi=300)
    plt.clf()
    plt.figure(2)
    plt.legend(loc='best')
    plt.subplots_adjust(left=0.12, right=0.995, top=0.990, bottom=0.105)
    # plt.title(f"{seq_name.replace('_', '-')}", fontdict=fontdict)
    plt.ylabel('MS-SSIM', fontdict=fontdict)
    plt.xlabel('bits per pixel (bpp)', fontdict=fontdict)
    plt.grid()
    #plt.show()
    plt.savefig(f'./figs/MSSSIM_{seq_name}.png', dpi=300)
    plt.cla()
    seq_idx += 1
    
    
ls = '-'
marker = 'o'
codec_idx = 0
for codec_name in codec_names:
    all_seq_bpps[codec_name] = all_seq_bpps[codec_name].tolist()
    all_seq_psnrs[codec_name] = all_seq_psnrs[codec_name].tolist()
    all_seq_msssims[codec_name] = all_seq_msssims[codec_name].tolist()
    all_seq_msssim_bpps[codec_name] = all_seq_msssim_bpps[codec_name].tolist()
    if codec_idx == 10:
        ls = '-.'
        marker = '^'
    plt.figure(1)
    plt.plot(all_seq_bpps[codec_name],all_seq_psnrs[codec_name],label=codec_names_t[codec_idx], marker=marker, ls=ls)
    #if codec_name in ['AV1', 'AV2']:
    #    codec_idx += 1
    #    continue
    plt.figure(2)
    if codec_name == 'HM' or codec_name == 'VTM':
        all_seq_msssim_bpps[codec_name] = all_seq_msssim_bpps[codec_name][:-2]
        all_seq_msssims[codec_name] = all_seq_msssims[codec_name][:-2]
    elif codec_name == 'DCVC-FM':
        all_seq_msssim_bpps[codec_name] = all_seq_msssim_bpps[codec_name][1:]
        all_seq_msssims[codec_name] = all_seq_msssims[codec_name][1:]
    plt.plot(all_seq_msssim_bpps[codec_name],all_seq_msssims[codec_name],label=codec_names_t[codec_idx], marker=marker, ls=ls)
    codec_idx += 1
    

plt.figure(1)
plt.legend(loc='best', prop={'size': 13})
plt.subplots_adjust(left=0.12, right=0.995, top=0.990, bottom=0.105)
#plt.legend(loc='best')
#plt.title('Average', fontdict=fontdict)
plt.ylabel('PSNR (dB)', fontdict=fontdict)
plt.xlabel('bits per pixel (bpp)', fontdict=fontdict)
plt.grid()
plt.savefig(f'./figs/PSNR_AVG.png', dpi=300)
plt.clf()
plt.figure(2)
plt.legend(loc='best', prop={'size': 13})
plt.subplots_adjust(left=0.12, right=0.995, top=0.990, bottom=0.105)
#plt.legend(loc='best')
#plt.title('Average', fontdict=fontdict)
plt.ylabel('MS-SSIM', fontdict=fontdict)
plt.xlabel('bits per pixel (bpp)', fontdict=fontdict)
plt.grid()
plt.savefig(f'./figs/MSSSIM_AVG.png', dpi=300)
#plt.show()
plt.cla()

json_str = json.dumps(all_seq_bpps, indent=4)
with open('all_seq_bpps_ivc.json', 'w') as json_file:
    json_file.write(json_str)

json_str = json.dumps(all_seq_psnrs, indent=4)
with open('all_seq_psnrs_ivc.json', 'w') as json_file:
    json_file.write(json_str)
    
json_str = json.dumps(all_seq_msssims, indent=4)
with open('all_seq_msssims_ivc.json', 'w') as json_file:
    json_file.write(json_str)
        